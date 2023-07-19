import requests
import xml.etree.ElementTree as ET
import urllib3
import warnings
from openpyxl import Workbook

# Jenkins服务器URL
jenkins_url = 'https://x.x.x.x:8080'

# 设置身份验证信息
username = 'jenkins_user'
password = 'jenkins_password'
auth = (username, password)

# 禁用不安全的 HTTPS 请求警告
warnings.filterwarnings("ignore", category=urllib3.exceptions.InsecureRequestWarning)

# 发送GET请求获取所有项目
api_url = f'{jenkins_url}/api/json'
response = requests.get(api_url, auth=auth, verify=False)
data = response.json()

# 解析JSON响应，获取项目名称
projects = [item['name'] for item in data['jobs']]

# 创建Excel工作簿和工作表
workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Git URLs'

# 设置表头
worksheet['A1'] = 'Project'
worksheet['B1'] = 'Git URL'

# 存储项目名称和Git地址
project_git_urls = []

# 遍历每个项目
row = 2
for project in projects:
    # 构建配置文件URL
    config_url = f'{jenkins_url}/job/{project}/config.xml'
    
    # 发送GET请求获取配置文件
    response = requests.get(config_url, auth=auth, verify=False)
    
    # 检查响应状态码并处理响应数据
    if response.status_code == 200:
        config_xml = response.text
        
        # 解析配置文件，提取Git地址
        tree = ET.ElementTree(ET.fromstring(config_xml))
        root = tree.getroot()
        
        # 检查是否存在scm元素
        scm_element = root.find('scm')
        if scm_element is not None:
            # 检查是否存在url元素
            git_url_element = scm_element.find('.//url')
            if git_url_element is not None:
                git_url = git_url_element.text
                
                # 将项目名称和Git地址存储到列表中
                project_git_urls.append((project, git_url))
                
                # 将数据写入Excel表格
                worksheet.cell(row=row, column=1, value=project)
                worksheet.cell(row=row, column=2, value=git_url)
                row += 1
    else:
        print(f"Failed to retrieve configuration file for project '{project}'. Status code: {response.status_code}")

# 保存Excel文件
workbook.save('git_urls.xlsx')
print("Git URLs saved to git_urls.xlsx")
